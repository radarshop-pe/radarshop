from flask import Blueprint, jsonify, request, send_file, session
from models import db, Product, Provider, Client, Sale, SaleDetail, Category, Inquiry, Seller
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func

inventory_bp = Blueprint('inventory', __name__)

# ── AUTH MIDDLEWARE ───────────────────────────────────────
@inventory_bp.before_request
def check_auth():
    """Protege todas las rutas /api/* — requiere sesión activa."""
    if request.method == 'OPTIONS':
        return  # permitir preflight CORS
    if 'user' not in session:
        return jsonify({'ok': False, 'message': 'No autorizado. Inicia sesión.'}), 401

# ── HELPERS ───────────────────────────────────────────────
def ok(data=None, msg='OK', code=200):
    return jsonify({'ok': True, 'message': msg, 'data': data}), code

def err(msg='Error', code=400):
    return jsonify({'ok': False, 'message': msg}), code

def style_sheet(ws, headers, rows, widths):
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        cell.fill  = PatternFill('solid', fgColor='FE7301')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# ── STATS / DASHBOARD ─────────────────────────────────────
@inventory_bp.route('/api/stats', methods=['GET'])
def get_stats():
    total_products = Product.query.count()
    low_stock = Product.query.filter(Product.stock_current <= Product.stock_min).count()
    total_clients = Client.query.count()

    products = Product.query.all()
    inventory_value = sum(p.cost_unit * p.stock_current for p in products)
    avg_margin = round(sum(p.margin_pct for p in products) / len(products), 1) if products else 0

    sales = Sale.query.all()
    total_sales = len(sales)
    total_income = sum(s.total_income for s in sales)
    total_profit = sum(s.total_profit for s in sales)
    avg_ticket = round(total_income / total_sales, 2) if total_sales else 0

    inquiries = Inquiry.query.all()
    total_inq = len(inquiries)
    converted = sum(1 for i in inquiries if i.result == 'Compró')
    conversion = round(converted / total_inq * 100, 1) if total_inq else 0

    # Top productos por ventas
    top_raw = db.session.query(
        SaleDetail.product_id,
        func.sum(SaleDetail.quantity).label('total_qty'),
        func.sum(SaleDetail.quantity * SaleDetail.price_at_sale).label('total_rev')
    ).group_by(SaleDetail.product_id)\
     .order_by(func.sum(SaleDetail.quantity).desc())\
     .limit(5).all()

    top_products = []
    for row in top_raw:
        p = Product.query.get(row.product_id)
        if p:
            top_products.append({
                'name': p.name,
                'qty': int(row.total_qty),
                'revenue': round(float(row.total_rev), 2)
            })

    # Ventas por canal
    channels_raw = db.session.query(
        Sale.channel, func.count(Sale.id)
    ).group_by(Sale.channel).all()
    channels = [{'channel': c, 'count': n} for c, n in channels_raw]

    return ok({
        'total_products': total_products,
        'low_stock': low_stock,
        'inventory_value': round(inventory_value, 2),
        'avg_margin': avg_margin,
        'total_clients': total_clients,
        'total_sales': total_sales,
        'total_income': round(total_income, 2),
        'total_profit': round(total_profit, 2),
        'avg_ticket': avg_ticket,
        'total_inquiries': total_inq,
        'conversion_rate': conversion,
        'top_products': top_products,
        'channels': channels
    })

# ── CATEGORÍAS ────────────────────────────────────────────
@inventory_bp.route('/api/categories', methods=['GET'])
def get_categories():
    cats = Category.query.order_by(Category.name).all()
    result = []
    for c in cats:
        d = c.to_dict()
        d['product_count'] = len(c.products)
        result.append(d)
    return ok(result)

@inventory_bp.route('/api/categories', methods=['POST'])
def add_category():
    d = request.json
    name = (d.get('name') or '').strip()
    if not name:
        return err('El nombre es requerido')
    if Category.query.filter_by(name=name).first():
        return err('Ya existe una categoría con ese nombre')
    cat = Category(name=name, description=d.get('description', ''))
    db.session.add(cat)
    db.session.commit()
    return ok(cat.to_dict(), 'Categoría creada', 201)

@inventory_bp.route('/api/categories/<int:cid>', methods=['PUT'])
def update_category(cid):
    cat = Category.query.get(cid)
    if not cat:
        return err('No encontrado', 404)
    d = request.json
    cat.name = d.get('name', cat.name)
    cat.description = d.get('description', cat.description)
    db.session.commit()
    return ok(cat.to_dict())

@inventory_bp.route('/api/categories/<int:cid>', methods=['DELETE'])
def delete_category(cid):
    cat = Category.query.get(cid)
    if not cat:
        return err('No encontrado', 404)
    if cat.products:
        return err('No se puede eliminar: tiene productos asociados')
    db.session.delete(cat)
    db.session.commit()
    return ok(msg='Categoría eliminada')

# ── VENDEDORES ────────────────────────────────────────────
@inventory_bp.route('/api/sellers', methods=['GET'])
def get_sellers():
    sellers = Seller.query.order_by(Seller.name).all()
    result = []
    for s in sellers:
        d = s.to_dict()
        d['total_sales'] = len(s.sales)
        result.append(d)
    return ok(result)

@inventory_bp.route('/api/sellers', methods=['POST'])
def create_seller():
    d = request.json
    if not d.get('name'):
        return err('El nombre es requerido')
    seller = Seller(
        name=d['name'].strip(),
        phone=d.get('phone', ''),
        status=d.get('status', 'Activo'),
        notes=d.get('notes', '')
    )
    db.session.add(seller)
    db.session.commit()
    return ok(seller.to_dict(), 'Vendedor creado', 201)

@inventory_bp.route('/api/sellers/<int:sid>', methods=['PUT'])
def update_seller(sid):
    s = Seller.query.get(sid)
    if not s:
        return err('No encontrado', 404)
    d = request.json
    s.name = d.get('name', s.name)
    s.phone = d.get('phone', s.phone)
    s.status = d.get('status', s.status)
    s.notes = d.get('notes', s.notes)
    db.session.commit()
    return ok(s.to_dict())

@inventory_bp.route('/api/sellers/<int:sid>', methods=['DELETE'])
def delete_seller(sid):
    s = Seller.query.get(sid)
    if not s:
        return err('No encontrado', 404)
    if s.sales:
        return err('No se puede eliminar: tiene ventas asociadas')
    db.session.delete(s)
    db.session.commit()
    return ok(msg='Vendedor eliminado')

# ── PRODUCTOS ─────────────────────────────────────────────
@inventory_bp.route('/api/products', methods=['GET'])
def get_products():
    ps = Product.query.order_by(Product.name).all()
    return ok([p.to_dict() for p in ps])

@inventory_bp.route('/api/products', methods=['POST'])
def add_product():
    d = request.json
    if not d.get('name') or not d.get('cost_unit') or not d.get('price_retail'):
        return err('Nombre, costo y precio de venta son requeridos')
    count = Product.query.count()
    sku = f"P{str(count + 1).zfill(3)}"
    while Product.query.get(sku):
        count += 1
        sku = f"P{str(count + 1).zfill(3)}"
    p = Product(
        id=sku,
        name=d['name'].strip(),
        color=d.get('color', ''),
        category_id=d.get('category_id') or None,
        cost_unit=float(d['cost_unit']),
        price_retail=float(d['price_retail']),
        price_wholesale=float(d.get('price_wholesale') or 0),
        stock_current=int(d.get('stock_current', 0)),
        stock_min=int(d.get('stock_min', 1)),
        location=d.get('location', ''),
        provider_id=d.get('provider_id') or None,
        status='Activo',
        notes=d.get('notes', '')
    )
    db.session.add(p)
    db.session.commit()
    return ok(p.to_dict(), 'Producto creado', 201)

@inventory_bp.route('/api/products/<pid>', methods=['PUT'])
def update_product(pid):
    p = Product.query.get(pid)
    if not p:
        return err('Producto no encontrado', 404)
    d = request.json
    p.name = d.get('name', p.name)
    p.color = d.get('color', p.color)
    p.category_id = d.get('category_id') or p.category_id
    p.cost_unit = float(d.get('cost_unit', p.cost_unit))
    p.price_retail = float(d.get('price_retail', p.price_retail))
    p.price_wholesale = float(d.get('price_wholesale') or p.price_wholesale or 0)
    p.stock_current = int(d.get('stock_current', p.stock_current))
    p.stock_min = int(d.get('stock_min', p.stock_min))
    p.location = d.get('location', p.location)
    p.provider_id = d.get('provider_id') or p.provider_id
    p.status = d.get('status', p.status)
    p.notes = d.get('notes', p.notes)
    db.session.commit()
    return ok(p.to_dict())

@inventory_bp.route('/api/products/<pid>', methods=['DELETE'])
def delete_product(pid):
    p = Product.query.get(pid)
    if not p:
        return err('No encontrado', 404)
    if p.sale_details:
        return err('No se puede eliminar: tiene ventas asociadas. Cambia su estado a Inactivo.')
    db.session.delete(p)
    db.session.commit()
    return ok(msg='Producto eliminado')

# ── CLIENTES ──────────────────────────────────────────────
@inventory_bp.route('/api/clients', methods=['GET'])
def get_clients():
    cs = Client.query.order_by(Client.name).all()
    return ok([c.to_dict() for c in cs])

@inventory_bp.route('/api/clients/search', methods=['GET'])
def search_client():
    q = request.args.get('q', '').strip()
    if not q:
        return ok([])
    cs = Client.query.filter(
        db.or_(
            Client.name.ilike(f'%{q}%'),
            Client.phone.ilike(f'%{q}%')
        )
    ).limit(10).all()
    return ok([c.to_dict() for c in cs])

@inventory_bp.route('/api/clients', methods=['POST'])
def add_client():
    d = request.json
    if not d.get('name'):
        return err('El nombre es requerido')
    count = Client.query.count()
    new_id = f"C{str(count + 1).zfill(3)}"
    while Client.query.get(new_id):
        count += 1
        new_id = f"C{str(count + 1).zfill(3)}"
    c = Client(
        id=new_id,
        name=d['name'].strip(),
        phone=d.get('phone', ''),
        district=d.get('district', ''),
        status=d.get('status', 'Activo'),
        notes=d.get('notes', '')
    )
    db.session.add(c)
    db.session.commit()
    return ok(c.to_dict(), 'Cliente registrado', 201)

@inventory_bp.route('/api/clients/<cid>', methods=['PUT'])
def update_client(cid):
    c = Client.query.get(cid)
    if not c:
        return err('Cliente no encontrado', 404)
    d = request.json
    c.name = d.get('name', c.name)
    c.phone = d.get('phone', c.phone)
    c.district = d.get('district', c.district)
    c.status = d.get('status', c.status)
    c.notes = d.get('notes', c.notes)
    db.session.commit()
    return ok(c.to_dict())

@inventory_bp.route('/api/clients/<cid>', methods=['DELETE'])
def delete_client(cid):
    c = Client.query.get(cid)
    if not c:
        return err('No encontrado', 404)
    if c.sales:
        return err('No se puede eliminar: tiene ventas asociadas')
    db.session.delete(c)
    db.session.commit()
    return ok(msg='Cliente eliminado')

# ── PROVEEDORES ───────────────────────────────────────────
@inventory_bp.route('/api/providers', methods=['GET'])
def get_providers():
    ps = Provider.query.order_by(Provider.name).all()
    return ok([p.to_dict() for p in ps])

@inventory_bp.route('/api/providers', methods=['POST'])
def add_provider():
    d = request.json
    if not d.get('name'):
        return err('El nombre es requerido')
    count = Provider.query.count()
    new_id = f"PR{str(count + 1).zfill(3)}"
    while Provider.query.get(new_id):
        count += 1
        new_id = f"PR{str(count + 1).zfill(3)}"
    p = Provider(
        id=new_id,
        name=d['name'].strip(),
        contact=d.get('contact', ''),
        phone=d.get('phone', ''),
        location=d.get('location', ''),
        type=d.get('type', 'Importador'),
        quality=int(d.get('quality', 3)),
        notes=d.get('notes', '')
    )
    db.session.add(p)
    db.session.commit()
    return ok(p.to_dict(), 'Proveedor guardado', 201)

@inventory_bp.route('/api/providers/<pid>', methods=['PUT'])
def update_provider(pid):
    p = Provider.query.get(pid)
    if not p:
        return err('No encontrado', 404)
    d = request.json
    p.name = d.get('name', p.name)
    p.contact = d.get('contact', p.contact)
    p.phone = d.get('phone', p.phone)
    p.location = d.get('location', p.location)
    p.type = d.get('type', p.type)
    p.quality = int(d.get('quality', p.quality))
    p.notes = d.get('notes', p.notes)
    db.session.commit()
    return ok(p.to_dict())

@inventory_bp.route('/api/providers/<pid>', methods=['DELETE'])
def delete_provider(pid):
    p = Provider.query.get(pid)
    if not p:
        return err('No encontrado', 404)
    if p.products:
        return err('No se puede eliminar: tiene productos asociados')
    db.session.delete(p)
    db.session.commit()
    return ok(msg='Proveedor eliminado')

# ── VENTAS ────────────────────────────────────────────────
@inventory_bp.route('/api/sales', methods=['GET'])
def get_sales():
    sales = Sale.query.order_by(Sale.date.desc()).all()
    return ok([s.to_dict() for s in sales])

@inventory_bp.route('/api/sales', methods=['POST'])
def create_sale():
    data = request.json
    try:
        # Vendedor: buscar por nombre o crear si no existe
        v_name = (data.get('seller_name') or 'Vendedor General').strip()
        seller = Seller.query.filter_by(name=v_name).first()
        if not seller:
            seller = Seller(name=v_name, status='Activo')
            db.session.add(seller)
            db.session.flush()

        # Cliente: buscar por nombre o crear si no existe
        c_name = (data.get('client_name') or 'Cliente General').strip()
        client = Client.query.filter_by(name=c_name).first()
        if not client:
            c_count = Client.query.count()
            new_cid = f"C{str(c_count + 1).zfill(3)}"
            while Client.query.get(new_cid):
                c_count += 1
                new_cid = f"C{str(c_count + 1).zfill(3)}"
            client = Client(id=new_cid, name=c_name, status='Activo')
            db.session.add(client)
            db.session.flush()

        delivery_fee = float(data.get('delivery_fee', 0))

        new_sale = Sale(
            id=f"V{datetime.now().strftime('%d%m%H%M%S')}",
            client_id=client.id,
            seller_id=seller.id,
            channel=data.get('channel', 'WhatsApp'),
            payment_method=data.get('payment_method', 'Efectivo'),
            delivery_type=data.get('delivery_type', 'Delivery'),
            delivery_cost=delivery_fee,
            notes=data.get('notes', ''),
            status='Completado'
        )
        db.session.add(new_sale)

        items = data.get('items', [])
        if not items:
            db.session.rollback()
            return err('La venta debe tener al menos un producto')

        # Validar stock primero
        for item in items:
            prod = Product.query.get(item.get('product_id'))
            if not prod:
                db.session.rollback()
                return err(f"Producto {item.get('product_id')} no encontrado")
            if prod.stock_current < int(item.get('quantity', 1)):
                db.session.rollback()
                return err(f"Stock insuficiente para '{prod.name}'. Disponible: {prod.stock_current}")

        # Registrar detalles y descontar stock
        for item in items:
            prod = Product.query.get(item['product_id'])
            qty = int(item.get('quantity', 1))
            price = float(item.get('price', prod.price_retail))
            detail = SaleDetail(
                sale_id=new_sale.id,
                product_id=prod.id,
                quantity=qty,
                price_at_sale=price,
                cost_at_sale=prod.cost_unit
            )
            prod.stock_current -= qty
            db.session.add(detail)

        db.session.commit()
        return ok({'id': new_sale.id}, 'Venta registrada', 201)

    except Exception as e:
        db.session.rollback()
        return err(str(e))

@inventory_bp.route('/api/sales/<sid>', methods=['DELETE'])
def delete_sale(sid):
    s = Sale.query.get(sid)
    if not s:
        return err('No encontrado', 404)
    # Revertir stock
    for detail in s.details:
        p = Product.query.get(detail.product_id)
        if p:
            p.stock_current += detail.quantity
    for detail in s.details:
        db.session.delete(detail)
    db.session.delete(s)
    db.session.commit()
    return ok(msg='Venta eliminada y stock revertido')

@inventory_bp.route('/api/sales/<sid>/status', methods=['PUT'])
def update_sale_status(sid):
    s = Sale.query.get(sid)
    if not s:
        return err('No encontrado', 404)
    d = request.json
    new_status = d.get('status')
    if new_status not in ['Completado', 'Cancelado']:
        return err('Estado inválido')
    if s.status == 'Completado' and new_status == 'Cancelado':
        for detail in s.details:
            p = Product.query.get(detail.product_id)
            if p:
                p.stock_current += detail.quantity
    elif s.status == 'Cancelado' and new_status == 'Completado':
        for detail in s.details:
            p = Product.query.get(detail.product_id)
            if p and p.stock_current >= detail.quantity:
                p.stock_current -= detail.quantity
            else:
                return err(f"Stock insuficiente para '{p.name}' al completar la venta")
    s.status = new_status
    db.session.commit()
    return ok(s.to_dict(), f"Estado actualizado a {new_status}")

# ── CONSULTAS ─────────────────────────────────────────────
@inventory_bp.route('/api/inquiries', methods=['GET'])
def get_inquiries():
    qs = Inquiry.query.order_by(Inquiry.date.desc()).all()
    return ok([q.to_dict() for q in qs])

@inventory_bp.route('/api/inquiries', methods=['POST'])
def add_inquiry():
    d = request.json
    if not d.get('product_name'):
        return err('El producto es requerido')
    inq = Inquiry(
        product_id=d.get('product_id') or None,
        product_name=d.get('product_name', ''),
        client_name=d.get('client_name', ''),
        channel=d.get('channel', ''),
        question=d.get('question', ''),
        result=d.get('result', 'Pendiente'),
        no_buy_reason=d.get('no_buy_reason', '')
    )
    db.session.add(inq)
    db.session.commit()
    return ok(inq.to_dict(), 'Consulta registrada', 201)

@inventory_bp.route('/api/inquiries/<int:iid>', methods=['DELETE'])
def delete_inquiry(iid):
    inq = Inquiry.query.get(iid)
    if not inq:
        return err('No encontrado', 404)
    db.session.delete(inq)
    db.session.commit()
    return ok(msg='Consulta eliminada')

# ── EXPORTAR EXCEL ────────────────────────────────────────
@inventory_bp.route('/api/export/excel', methods=['GET'])
def export_excel():
    wb = openpyxl.Workbook()

    # Inventario
    ws1 = wb.active
    ws1.title = 'Inventario'
    products = Product.query.order_by(Product.name).all()
    style_sheet(ws1,
        ['SKU', 'Nombre', 'Categoría', 'Color', 'Costo', 'P. Venta', 'P. Mayor',
         'Margen %', 'Stock', 'Stock Min', 'Alerta', 'Proveedor', 'Ubicación', 'Estado'],
        [(p.id, p.name,
          p.category_obj.name if p.category_obj else '',
          p.color or '', p.cost_unit, p.price_retail,
          p.price_wholesale or 0, p.margin_pct,
          p.stock_current, p.stock_min,
          '⚠' if p.stock_alert else 'OK',
          p.provider.name if p.provider else '',
          p.location or '', p.status)
         for p in products],
        [8, 28, 14, 10, 9, 9, 9, 9, 7, 8, 7, 18, 12, 10]
    )

    # Ventas
    ws2 = wb.create_sheet('Ventas')
    sales = Sale.query.order_by(Sale.date.desc()).all()
    rows_v = []
    for s in sales:
        for d in s.details:
            rows_v.append((
                s.id,
                s.date.strftime('%d/%m/%Y') if s.date else '',
                s.client.name if s.client else '',
                s.seller.name if s.seller else '',
                d.product.name if d.product else '',
                d.quantity, d.price_at_sale,
                round(d.quantity * d.price_at_sale, 2),
                s.channel, s.payment_method,
                s.delivery_type, s.delivery_cost,
                s.total_income, s.total_profit, s.status
            ))
    style_sheet(ws2,
        ['ID Venta', 'Fecha', 'Cliente', 'Vendedor', 'Producto', 'Cantidad',
         'Precio Unit.', 'Subtotal', 'Canal', 'Pago', 'Entrega',
         'Delivery', 'Total Ingreso', 'Ganancia', 'Estado'],
        rows_v,
        [12, 11, 18, 18, 22, 8, 11, 11, 16, 12, 12, 9, 13, 11, 11]
    )

    # Clientes
    ws3 = wb.create_sheet('Clientes')
    clients = Client.query.order_by(Client.name).all()
    style_sheet(ws3,
        ['ID', 'Nombre', 'Teléfono', 'Distrito', 'Estado',
         'Total Pedidos', 'Total Gastado', 'Fecha Registro'],
        [(c.id, c.name, c.phone or '', c.district or '', c.status,
          c.total_orders, c.total_spent,
          c.created_at.strftime('%d/%m/%Y') if c.created_at else '')
         for c in clients],
        [8, 24, 14, 16, 10, 13, 14, 14]
    )

    # Consultas
    ws4 = wb.create_sheet('Consultas')
    inquiries = Inquiry.query.order_by(Inquiry.date.desc()).all()
    style_sheet(ws4,
        ['Fecha', 'Producto', 'Cliente', 'Canal', 'Pregunta', 'Resultado', 'Motivo No Compra'],
        [(q.date.strftime('%d/%m/%Y') if q.date else '',
          q.product_name, q.client_name or '',
          q.channel or '', q.question or '',
          q.result, q.no_buy_reason or '')
         for q in inquiries],
        [11, 22, 16, 16, 30, 11, 24]
    )

    # Proveedores
    ws5 = wb.create_sheet('Proveedores')
    providers = Provider.query.order_by(Provider.name).all()
    style_sheet(ws5,
        ['ID', 'Nombre', 'Contacto', 'Teléfono', 'Ubicación', 'Tipo', 'Calidad', 'Notas'],
        [(p.id, p.name, p.contact or '', p.phone or '',
          p.location or '', p.type or '', p.quality, p.notes or '')
         for p in providers],
        [8, 22, 16, 14, 18, 13, 9, 28]
    )

    # Vendedores
    ws6 = wb.create_sheet('Vendedores')
    sellers = Seller.query.order_by(Seller.name).all()
    style_sheet(ws6,
        ['ID', 'Nombre', 'Teléfono', 'Estado', 'Total Ventas', 'Notas'],
        [(s.id, s.name, s.phone or '', s.status, len(s.sales), s.notes or '')
         for s in sellers],
        [8, 24, 14, 10, 12, 28]
    )

    # Categorías
    ws7 = wb.create_sheet('Categorías')
    cats = Category.query.order_by(Category.name).all()
    style_sheet(ws7,
        ['ID', 'Nombre', 'Descripción', 'N° Productos'],
        [(c.id, c.name, c.description or '', len(c.products)) for c in cats],
        [6, 22, 30, 13]
    )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"Radarshop_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )